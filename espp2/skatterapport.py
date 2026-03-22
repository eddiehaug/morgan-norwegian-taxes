"""
Generate Skattemeldingen reports from ESPP2 tax results.

Provides:
  - add_skattemeldingen_sheet(): Adds a "Skattemeldingen {year}" sheet to an
    existing openpyxl workbook (used by the web app and CLI xlsx output).
  - generate_skatterapport(): Legacy HTML report generator (kept for reference).
"""

from decimal import Decimal
from typing import Optional
import io


_COMPANY_NAMES = {
    "GOOG": "Alphabet Inc.",
    "GOOGL": "Alphabet Inc.",
    "CSCO": "Cisco Systems Inc.",
    "MSFT": "Microsoft Corporation",
    "AAPL": "Apple Inc.",
    "AMZN": "Amazon.com Inc.",
    "NVDA": "NVIDIA Corporation",
    "META": "Meta Platforms Inc.",
}


def _company_name(symbol: str) -> str:
    return _COMPANY_NAMES.get(symbol.upper(), symbol)


def _fmt(value, decimals=0) -> str:
    """Format a number with Norwegian thousand-separator (space)."""
    if decimals == 0:
        return f"{int(round(value)):,}".replace(",", "\u202f")
    return f"{float(value):,.{decimals}f}".replace(",", "\u202f")


def generate_skatterapport(
    result,
    year: int,
    broker_display: str = "Morgan Stanley",
    account_id: str = "",
    output_path: Optional[str] = None,
) -> str:
    """Generate HTML Skattemeldingen report.

    Args:
        result: TaxReportReturn from espp2 main.do_taxes()
        year: Tax year (e.g. 2025)
        broker_display: Human-readable broker name for the report
        account_id: Morgan Stanley account number (optional)
        output_path: Write HTML to this file path (optional)

    Returns:
        HTML string
    """
    summary = result.summary
    report = result.report

    sections_html = ""

    for fs in summary.foreignshares:
        # Get USD wealth from EOY balance items
        eoy_items = report.eoy_balance.get(year, [])
        symbol_items = [item for item in eoy_items if item.symbol == fs.symbol]
        wealth_usd = sum(item.amount.value for item in symbol_items)
        wealth_nok = fs.wealth

        if wealth_usd and wealth_usd != 0:
            exchange_rate = (Decimal(str(wealth_nok)) / Decimal(str(wealth_usd))).quantize(Decimal("0.0001"))
        else:
            exchange_rate = Decimal("1")

        # Credit deduction for this symbol
        cd = next((c for c in summary.credit_deduction if c.symbol == fs.symbol), None)

        # Gain or loss
        gain_val = int(round(fs.taxable_gain))
        if gain_val >= 0:
            gain_row = f'<tr><td>Skattepliktig gevinst</td><td class="val">{_fmt(gain_val)} NOK</td></tr>'
            loss_row = ""
        else:
            gain_row = '<tr><td>Skattepliktig gevinst</td><td class="val empty">—</td></tr>'
            loss_row = f'<tr><td>Fradragsberettiget tap</td><td class="val loss">{_fmt(abs(gain_val))} NOK</td></tr>'

        account_row = (
            f'<tr><td>Kontonummer eller annen identifikator</td><td class="val">{account_id}</td></tr>'
            if account_id
            else ""
        )

        sections_html += f"""
        <section>
            <h2>Utenlandske aksjer – {fs.symbol} ({year})</h2>
            <table>
                <tr><th>Felt (Skattemeldingen)</th><th>Verdi</th></tr>
                <tr><td>Land</td><td class="val">{fs.country}</td></tr>
                <tr><td>Kontofører/bank</td><td class="val">{broker_display}</td></tr>
                {account_row}
                <tr><td>Navn på aksjeselskap</td><td class="val">{_company_name(fs.symbol)}</td></tr>
                <tr><td>ISIN</td><td class="val mono">{fs.isin}</td></tr>
                <tr><td>Antall aksjer per 31. desember</td><td class="val">{fs.shares:.4f}</td></tr>
                <tr class="sub"><td colspan="2">Formue (markedsverdi per 31. desember)</td></tr>
                <tr><td>&nbsp;&nbsp;&nbsp;Valuta</td><td class="val">USD – Amerikansk dollar</td></tr>
                <tr><td>&nbsp;&nbsp;&nbsp;Beløp (USD)</td><td class="val">{_fmt(wealth_usd)}</td></tr>
                <tr><td>&nbsp;&nbsp;&nbsp;Beløp i NOK</td><td class="val">{_fmt(wealth_nok)}</td></tr>
                <tr><td>&nbsp;&nbsp;&nbsp;Valutakurs (NOK/USD)</td><td class="val">{exchange_rate}</td></tr>
                <tr class="sub"><td colspan="2">Inntekt og fradrag</td></tr>
                <tr><td>Skattepliktig utbytte</td><td class="val">{_fmt(fs.dividend)} NOK</td></tr>
                <tr><td>&nbsp;&nbsp;&nbsp;Valuta</td><td class="val">NOK – Norsk krone</td></tr>
                {gain_row}
                {loss_row}
                <tr><td>Anvendt skjerming</td><td class="val">{_fmt(fs.tax_deduction_used)} NOK</td></tr>
            </table>
        </section>"""

        if cd:
            sections_html += f"""
        <section>
            <h2>Kreditfradrag – dobbeltbeskatning ({fs.symbol})</h2>
            <p class="hint">Fradrag for skatt betalt i USA på aksjeutbytte (kildeskatt 15%).</p>
            <table>
                <tr><th>Felt (Skattemeldingen)</th><th>Verdi</th></tr>
                <tr><td>Land</td><td class="val">{cd.country}</td></tr>
                <tr><td>Inntektsskatt betalt i utlandet</td><td class="val">{_fmt(cd.income_tax)} NOK</td></tr>
                <tr><td>Brutto aksjeutbytte</td><td class="val">{_fmt(cd.gross_share_dividend)} NOK</td></tr>
                <tr><td>Herav skatt på brutto aksjeutbytte</td><td class="val">{_fmt(cd.tax_on_gross_share_dividend)} NOK</td></tr>
            </table>
        </section>"""

    # Currency gain/loss — only the non-aggregated part needs separate reporting
    fx_nonagg = int(round(summary.cashsummary.gain))
    fx_agg    = int(round(summary.cashsummary.gain_aggregated))

    # Per-wire detail rows
    wire_rows_html = ""
    for t in summary.cashsummary.transfers:
        wire_gain = int(round(float(t.gain))) + int(round(float(t.aggregated_gain)))
        is_agg = int(round(float(t.aggregated_gain))) != 0
        sign = "+" if wire_gain >= 0 else ""
        behandling = "Type A — i aksjegevinst/-tap" if is_agg else "Type B — Finans → Valuta"
        cls = "gain" if wire_gain >= 0 else "loss"
        wire_rows_html += (
            f'<tr><td>{t.date}</td>'
            f'<td class="val">{_fmt(int(round(float(t.amount_sent))))} NOK</td>'
            f'<td class="val">{_fmt(int(round(float(t.amount_received))))} NOK</td>'
            f'<td class="val {cls}">{sign}{_fmt(abs(wire_gain))} NOK</td>'
            f'<td>{behandling}</td></tr>'
        )

    agg_note = (
        f'<span class="loss">{_fmt(abs(fx_agg))} NOK tap</span>' if fx_agg < 0 else
        f'<span class="gain">{_fmt(fx_agg)} NOK gevinst</span>'
    )
    nonagg_note = (
        f'<span class="loss">{_fmt(abs(fx_nonagg))} NOK tap</span>' if fx_nonagg < 0 else
        (f'<span class="gain">{_fmt(fx_nonagg)} NOK gevinst</span>' if fx_nonagg > 0 else
         '<span class="empty">0 NOK — ingen separat rapportering nødvendig</span>')
    )

    sections_html += f"""
        <section>
            <h2>Valutakursgevinst/-tap — USD-konto (Morgan Stanley)</h2>
            <p class="hint">
                Valutaresultatet er delt i to typer basert på om pengene ble overført innen 14 dager etter salget.
            </p>
            <table>
                <tr><th colspan="2">Oppsummering</th></tr>
                <tr>
                    <td><strong>Type A — Aggregert (inkludert i aksjetap/-gevinst)</strong><br>
                        <span class="hint">USD fra aksjesalg overført innen 14 dager. Allerede inkludert i
                        Fradragsberettiget tap / Skattepliktig gevinst ovenfor. <em>Rapporter ikke separat.</em></span>
                    </td>
                    <td class="val">{agg_note}</td>
                </tr>
                <tr>
                    <td><strong>Type B — Separat valutagevinst/-tap</strong><br>
                        <span class="hint">USD fra utbytte eller salg overført etter 14 dager.
                        Rapporter under <em>Finans → Valuta</em> i Skattemeldingen.</span>
                    </td>
                    <td class="val">{nonagg_note}</td>
                </tr>
            </table>
            <br>
            <table>
                <tr>
                    <th>Dato</th><th>Kostpris (NOK)</th><th>Mottatt (NOK)</th>
                    <th>Gevinst/Tap</th><th>Behandling</th>
                </tr>
                {wire_rows_html}
            </table>
            <p class="hint">Kostpris = USD-beløp × Norges Bank-kurs på salgsdatoen. Gevinst/tap = Mottatt − Kostpris.</p>
        </section>"""

    # USD cash balance
    remaining = summary.cashsummary.remaining_cash
    remaining_usd = float(remaining.value)
    remaining_nok = int(round(remaining.nok_value)) if remaining.nok_value else 0
    sections_html += f"""
        <section>
            <h2>USD-kontobeholdning per 31. desember {year}</h2>
            <p class="hint">USD-saldo på Morgan Stanley-konto er skattepliktig som formue.</p>
            <table>
                <tr><th>Felt (Skattemeldingen)</th><th>Verdi</th></tr>
                <tr><td>Beholdning (USD)</td><td class="val">{remaining_usd:.2f} USD</td></tr>
                <tr><td>Formue (NOK)</td><td class="val">{_fmt(remaining_nok)} NOK</td></tr>
            </table>
        </section>"""

    html = _wrap_html(sections_html, year)

    if output_path:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html)

    return html


def _wrap_html(body: str, year: int) -> str:
    return f"""<!DOCTYPE html>
<html lang="no">
<head>
<meta charset="UTF-8">
<title>Skattemeldingen – Utenlandske aksjer {year}</title>
<style>
  body {{
    font-family: Arial, Helvetica, sans-serif;
    max-width: 860px;
    margin: 40px auto;
    color: #222;
    font-size: 14px;
    line-height: 1.5;
  }}
  h1 {{
    color: #1a3c6e;
    border-bottom: 3px solid #1a3c6e;
    padding-bottom: 10px;
    font-size: 22px;
  }}
  h2 {{
    color: #1a3c6e;
    margin-top: 32px;
    margin-bottom: 6px;
    background: #eef4fb;
    padding: 8px 14px;
    border-left: 5px solid #1a3c6e;
    font-size: 15px;
  }}
  section {{ margin-bottom: 10px; }}
  table {{
    width: 100%;
    border-collapse: collapse;
    margin-top: 4px;
  }}
  th {{
    background: #1a3c6e;
    color: white;
    text-align: left;
    padding: 7px 12px;
    font-size: 13px;
  }}
  td {{
    padding: 6px 12px;
    border-bottom: 1px solid #e8e8e8;
    font-size: 13px;
  }}
  tr:hover td {{ background: #f5f8ff; }}
  .val {{
    font-weight: bold;
    text-align: right;
    font-variant-numeric: tabular-nums;
  }}
  .mono {{ font-family: monospace; font-size: 13px; }}
  .empty {{ color: #999; font-weight: normal; }}
  .gain {{ color: #1a6e3c; }}
  .loss {{ color: #8b0000; }}
  .sub td {{
    background: #f0f4fa;
    font-style: italic;
    color: #555;
    font-size: 12px;
  }}
  .hint {{
    color: #666;
    margin: 2px 0 8px 0;
    font-size: 12px;
  }}
  .footer {{
    margin-top: 48px;
    padding-top: 12px;
    border-top: 1px solid #ccc;
    font-size: 11px;
    color: #999;
  }}
  @media print {{
    body {{ margin: 15mm; }}
    section {{ page-break-inside: avoid; }}
    h2 {{ background: none !important; border-left-color: #1a3c6e; }}
  }}
</style>
</head>
<body>
<h1>Skattemeldingen – Utenlandske aksjer {year}</h1>
<p>Bruk verdiene nedenfor til å fylle ut skattemeldingen på <strong>skatteetaten.no</strong>.<br>
Alle beløp er i NOK med mindre annet er angitt.</p>
{body}
</body>
</html>"""


# ---------------------------------------------------------------------------
# xlsx Skattemeldingen sheet
# ---------------------------------------------------------------------------

def add_skattemeldingen_sheet(workbook, result, year: int, account_id: str = "") -> None:
    """
    Add a 'Skattemeldingen {year}' sheet to an existing openpyxl workbook.

    The sheet lists every value that must be entered into the Skattemeldingen
    form on skatteetaten.no, organized by section.

    Args:
        workbook: openpyxl.Workbook object (modified in-place)
        result: TaxReportReturn from do_taxes()
        year: Tax year (e.g. 2024)
        account_id: Morgan Stanley account number (e.g. "MS-627432-82")
    """
    from openpyxl import styles
    from openpyxl.styles import Font, PatternFill, Alignment, numbers

    PRIMARY = "003178"
    PRIMARY_TEXT = "FFFFFF"
    SECTION_BG = "E8EEF8"
    LABEL_FONT = Font(name="Calibri", size=11)
    VALUE_FONT = Font(name="Calibri", size=11, bold=True)
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=PRIMARY_TEXT)
    SECTION_FONT = Font(name="Calibri", size=11, bold=True, color=PRIMARY)
    HEADER_FILL = PatternFill("solid", fgColor=PRIMARY)
    SECTION_FILL = PatternFill("solid", fgColor=SECTION_BG)
    RIGHT = Alignment(horizontal="right")
    LEFT = Alignment(horizontal="left")

    sheet_name = f"Skattemeldingen {year}"
    # Remove existing sheet with same name if present
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    ws = workbook.create_sheet(title=sheet_name)

    # Set column widths
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 46

    summary = result.summary
    report = result.report

    row = 1

    def _write_header(label):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = HEADER_FONT
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.cell(row=row, column=2).fill = HEADER_FILL
        row += 1

    def _write_section(label):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.cell(row=row, column=2).fill = SECTION_FILL
        row += 1

    def _write_row(label, value, note=""):
        nonlocal row
        c_label = ws.cell(row=row, column=1, value=label)
        c_label.font = LABEL_FONT
        c_label.alignment = LEFT
        c_val = ws.cell(row=row, column=2, value=value)
        c_val.font = VALUE_FONT
        c_val.alignment = RIGHT
        if note:
            ws.cell(row=row, column=3, value=note).font = Font(name="Calibri", size=9, italic=True, color="666666")
        row += 1

    def _blank():
        nonlocal row
        row += 1

    _write_header(f"Skattemeldingen {year} — Verdier for skatteetaten.no")
    _blank()

    for fs in summary.foreignshares:
        # --- Wealth ---
        eoy_items = report.eoy_balance.get(year, [])
        symbol_items = [item for item in eoy_items if item.symbol == fs.symbol]
        wealth_usd = float(sum(item.amount.value for item in symbol_items))
        wealth_nok = int(round(float(fs.wealth)))

        if wealth_usd and wealth_usd != 0:
            exchange_rate = round(wealth_nok / wealth_usd, 4)
        else:
            exchange_rate = 1.0

        cd = next((c for c in summary.credit_deduction if c.symbol == fs.symbol), None)
        gain_val = int(round(float(fs.taxable_gain)))

        # ---- Utenlandske aksjer ----
        _write_section(f"Utenlandske aksjer — {fs.symbol}")
        _write_row("Land", fs.country)
        _write_row("Kontofører/bank", "Morgan Stanley")
        if account_id:
            _write_row("Kontonummer eller annen identifikator", account_id)
        _write_row("Navn på aksjeselskap", _company_name(fs.symbol))
        _write_row("ISIN", fs.isin)
        _write_row("Antall aksjer per 31. desember", float(round(float(fs.shares), 4)))
        _blank()
        _write_section("Formue (markedsverdi per 31. desember)")
        _write_row("Valuta", "USD – Amerikansk dollar")
        _write_row("Beløp (USD)", round(wealth_usd, 2))
        _write_row("Beløp i NOK", wealth_nok, "Oppgi dette beløpet i Skattemeldingen")
        _write_row("Valutakurs (NOK/USD)", exchange_rate)
        _blank()
        _write_section("Inntekt og fradrag")
        _write_row("Skattepliktig utbytte (NOK)", int(round(float(fs.dividend))),
                   "Oppgi i NOK – Norsk krone")
        if gain_val >= 0:
            _write_row("Skattepliktig gevinst (NOK)", gain_val,
                       "Inkluderer gevinst fra aksjesalg og valutakursgevinst på salgsbeløp (Type A)")
            _write_row("Fradragsberettiget tap (NOK)", "—")
        else:
            _write_row("Skattepliktig gevinst (NOK)", "—")
            _write_row("Fradragsberettiget tap (NOK)", abs(gain_val),
                       "Inkluderer tap fra aksjesalg og valutatap på salgsbeløp (Type A) — se Valutaberegning nedenfor")
        _write_row("Anvendt skjerming (NOK)", int(round(float(fs.tax_deduction_used))))
        _blank()

        if cd:
            # ---- Kreditfradrag ----
            _write_section("Kreditfradrag — dobbeltbeskatning")
            _write_row("Land", cd.country)
            _write_row("Inntektsskatt betalt i utlandet (NOK)", int(round(float(cd.income_tax))))
            _write_row("Brutto aksjeutbytte (NOK)", int(round(float(cd.gross_share_dividend))))
            _write_row("Herav skatt på brutto aksjeutbytte (NOK)",
                       int(round(float(cd.tax_on_gross_share_dividend))))
            _blank()

    # ---- Valuta — beregningsdetaljer og rapportering ----
    cashsummary = summary.cashsummary
    fx_nonagg = int(round(float(cashsummary.gain)))            # separate Valutagevinst/-tap
    fx_agg    = int(round(float(cashsummary.gain_aggregated))) # already in share gain above

    _write_section("Valutakursgevinst/-tap — USD-konto (Morgan Stanley)")

    NOTE_FONT  = Font(name="Calibri", size=9, italic=True, color="666666")
    WARN_FILL  = PatternFill("solid", fgColor="FFF3CD")
    GREEN_FILL = PatternFill("solid", fgColor="D6F0E0")

    def _write_info(text, fill=None):
        nonlocal row
        c = ws.cell(row=row, column=1, value=text)
        c.font = NOTE_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        if fill:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = fill
        row += 1

    # Contextual summary rows
    _write_info(
        "Valutaberegning: for hvert bankuttak sammenlignes kostpris (Norges Bank-kurs på salgsdato) "
        "med faktisk mottatt NOK. Gevinst/tap deles i to typer — se forklaring nedenfor.",
        fill=PatternFill("solid", fgColor="EEF4FB"),
    )
    _blank()

    # --- Aggregated FX (already in share gain) ---
    _write_section("  Type A — Aggregert valutagevinst/-tap (inkludert i aksjetap/-gevinst ovenfor)")
    _write_info(
        "USD fra aksjesalg som ble overført til norsk bank innen 14 dager er behandlet som én "
        "transaksjon med salget. Valutaresultatet er allerede inkludert i "
        "«Fradragsberettiget tap» / «Skattepliktig gevinst» — rapporter IKKE dette separat.",
        fill=WARN_FILL,
    )
    agg_label = "Aggregert valutagevinst (NOK)" if fx_agg >= 0 else "Aggregert valutatap (NOK)"
    _write_row(agg_label, abs(fx_agg),
               "Allerede inkludert i Fradragsberettiget tap / Skattepliktig gevinst ovenfor")
    _blank()

    # --- Non-aggregated FX (separate reporting) ---
    _write_section("  Type B — Separat valutagevinst/-tap (rapporter under Finans → Valuta)")
    if fx_nonagg == 0:
        _write_info(
            "Ingen separat valutagevinst/-tap å rapportere. "
            "Alle bankoverføringer ble behandlet som Type A (aggregert med aksjesalg).",
            fill=GREEN_FILL,
        )
        _write_row("Separat Valutagevinst/-tap (NOK)", 0,
                   "Ingen rapportering nødvendig under Finans → Valuta")
    else:
        _write_info(
            "USD fra utbytte eller salg overført etter 14 dager behandles som separat valutagevinst/-tap. "
            "Rapporter dette under Finans → Valuta i Skattemeldingen.",
            fill=WARN_FILL,
        )
        nonagg_label = "Separat valutagevinst (NOK)" if fx_nonagg >= 0 else "Separat valutatap (NOK)"
        _write_row(nonagg_label, abs(fx_nonagg),
                   "Oppgi under Finans → Valuta i Skattemeldingen")
    _blank()

    # --- Per-wire breakdown table ---
    transfers = cashsummary.transfers
    if transfers:
        _write_section("  Detaljert oversikt — bankoverføringer")

        TBL_HDR_FONT = Font(name="Calibri", size=10, bold=True, color=PRIMARY_TEXT)
        TBL_HDR_FILL = PatternFill("solid", fgColor="4472C4")
        TBL_FONT     = Font(name="Calibri", size=10)
        TBL_FONT_B   = Font(name="Calibri", size=10, bold=True)
        TBL_ALT_FILL = PatternFill("solid", fgColor="F2F7FF")

        headers = ["Dato", "Kostpris (NOK)", "Mottatt (NOK)", "Gevinst/Tap (NOK)", "Behandling"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = TBL_HDR_FONT
            c.fill = TBL_HDR_FILL
            c.alignment = Alignment(horizontal="center")
        row += 1

        for i, t in enumerate(transfers):
            fill = TBL_ALT_FILL if i % 2 == 1 else None
            # Total gain for this wire = gain + aggregated_gain
            wire_gain = int(round(float(t.gain))) + int(round(float(t.aggregated_gain)))
            is_agg = int(round(float(t.aggregated_gain))) != 0

            cells_data = [
                (str(t.date), "left"),
                (int(round(float(t.amount_sent))),   "right"),
                (int(round(float(t.amount_received))), "right"),
                (wire_gain, "right"),
                ("Type A — Inkludert i aksjegevinst/-tap" if is_agg
                 else "Type B — Rapporter under Finans → Valuta", "left"),
            ]
            for col, (val, align) in enumerate(cells_data, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = TBL_FONT
                c.alignment = Alignment(horizontal=align)
                if fill:
                    c.fill = fill
            row += 1

        # Totals row
        TOTAL_FONT = Font(name="Calibri", size=10, bold=True)
        TOTAL_FILL = PatternFill("solid", fgColor="D0D9F0")
        total_sent     = int(round(sum(float(t.amount_sent)     for t in transfers)))
        total_received = int(round(sum(float(t.amount_received) for t in transfers)))
        total_gain_all = fx_agg + fx_nonagg
        totals = [
            ("Totalt", "left"),
            (total_sent,     "right"),
            (total_received, "right"),
            (total_gain_all, "right"),
            ("", "left"),
        ]
        for col, (val, align) in enumerate(totals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = TOTAL_FONT
            c.fill = TOTAL_FILL
            c.alignment = Alignment(horizontal=align)
        row += 1

    _blank()

    # ---- USD-kontobeholdning ----
    _write_section(f"USD-kontobeholdning per 31. desember {year}")
    remaining = summary.cashsummary.remaining_cash
    remaining_usd = round(float(remaining.value), 2)
    remaining_nok = int(round(float(remaining.nok_value))) if remaining.nok_value else 0
    _write_row("Beholdning (USD)", remaining_usd)
    _write_row("Formue (NOK)", remaining_nok, "Oppgi som formue i Skattemeldingen")

    _blank()
    ws.cell(row=row, column=1, value=f"Generert av TaxLedger Pro · Inntektsår {year}").font = \
        Font(name="Calibri", size=9, italic=True, color="999999")


def build_xlsx_with_skatterapport(excel_bytes: bytes, result, year: int, account_id: str = "") -> bytes:
    """
    Take the raw xlsx bytes from do_taxes(), open as workbook, add the
    Skattemeldingen sheet, and return the modified xlsx as bytes.
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    add_skattemeldingen_sheet(wb, result, year, account_id=account_id)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
